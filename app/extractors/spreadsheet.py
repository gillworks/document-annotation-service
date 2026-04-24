from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.extractors.base import (
    MAX_SAMPLE_ROWS,
    ExtractionError,
    ExtractionPayload,
    stringify_cell,
    truncate_text,
)


def extract_xlsx(path: Path) -> ExtractionPayload:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_payloads = []
        text_parts = []
        warnings = []

        for worksheet in workbook.worksheets:
            sample_rows: list[list[Any]] = []
            for row in worksheet.iter_rows(max_row=MAX_SAMPLE_ROWS, values_only=True):
                sample_rows.append([stringify_cell(value) for value in row])

            non_empty_rows = [row for row in sample_rows if any(cell for cell in row)]
            headers = first_non_empty_row(non_empty_rows)
            sheet_payload = {
                "name": worksheet.title,
                "row_count": worksheet.max_row or 0,
                "column_count": worksheet.max_column or 0,
                "headers": headers,
                "sample_rows": non_empty_rows,
            }
            sheet_payloads.append(sheet_payload)

            text_parts.append(render_sheet_text(sheet_payload))
            if (worksheet.max_row or 0) > MAX_SAMPLE_ROWS:
                warnings.append(
                    f"sheet {worksheet.title!r} sampled first {MAX_SAMPLE_ROWS} rows"
                )

        full_text, truncated = truncate_text("\n\n".join(text_parts))
        if truncated:
            warnings.append("extraction text truncated to 60000 characters")

        return ExtractionPayload(
            source_type="xlsx",
            text=full_text,
            sheets=sheet_payloads,
            metadata={
                "page_count": None,
                "sheet_count": len(workbook.worksheets),
                "sheet_names": workbook.sheetnames,
                "has_tables": any(
                    sheet["row_count"] > 1 and sheet["column_count"] > 1
                    for sheet in sheet_payloads
                ),
            },
            warnings=warnings,
        )
    except ExtractionError:
        raise
    except Exception as exc:
        raise ExtractionError(
            "SPREADSHEET_PARSE_FAILED",
            f"XLSX extraction failed: {exc}",
        ) from exc


def first_non_empty_row(rows: list[list[str]]) -> list[str]:
    for row in rows:
        if any(cell for cell in row):
            return row
    return []


def render_sheet_text(sheet: dict[str, Any]) -> str:
    lines = [
        f"Sheet: {sheet['name']}",
        f"Rows: {sheet['row_count']}, Columns: {sheet['column_count']}",
    ]
    if sheet["headers"]:
        lines.append("Headers: " + ", ".join(sheet["headers"]))
    for row in sheet["sample_rows"]:
        if any(row):
            lines.append(" | ".join(row))
    return "\n".join(lines)
